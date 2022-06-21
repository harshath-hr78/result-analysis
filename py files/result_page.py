from distutils.log import error
from selenium import webdriver
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook,cell
import time
import os
import pyautogui
import cv2 as cv
import pytesseract
import pandas as pd


# install all these as pip install filename, and pip install opencv-python.

option = webdriver.ChromeOptions()
option.add_argument("-incognito")
option.add_experimental_option("excludeSwitches", ['enable-automation'])
option.add_experimental_option("detach",True)

#add your chrome driver installation path
browser = webdriver.Chrome(executable_path=r'C:\Program Files (x86)\chromedriver.exe', options=option)

def fillLoginpage(usn,ite):

    browser.get(r"https://results.vtu.ac.in/FMEcbcs22/resultpage.php")

    #getting hold of usn and captcha input fields.

    testbox = browser.find_element_by_xpath("/html/body/div[2]/div[1]/div[2]/div/div[2]/form/div/div[2]/div[1]/div/input")
    captchabox = browser.find_element_by_xpath("/html/body/div[2]/div[1]/div[2]/div/div[2]/form/div/div[2]/div[2]/div[1]/input")

    #start with the image capta recognition procedure
    time.sleep(2)
    myScreenshot = pyautogui.screenshot(region=(970, 405, 170, 80))  #region=(horizontal pos, vertical pos, vertical ratio, horizontal ratio)
    myScreenshot.save(r'C:\Users\harsh\Desktop\result_analysis\pics\screenshot.png') #change according to your dir.

    os.chdir(r"C:\Users\harsh\Desktop\result_analysis\pics")
    img = cv.imread('screenshot.png',0)
    ret,thresh = cv.threshold(img,103,150,cv.THRESH_TOZERO_INV)
    cv.imshow('Binary Threshold', thresh)
    # Using cv2.imwrite() method
    # Saving the image
    os.chdir(r'C:\Users\harsh\Desktop\result_analysis\pics')
    cv.imwrite("thresh_img.png", thresh)

    time.sleep(1)
    #os.system('"wsl tesseract thresh_img.jpg result"') #tesseract is ocr function for image to text
    img2 = cv.imread('thresh_img.png',0)
    #install tesseract from https://github.com/UB-Mannheim/tesseract/wiki choose 64-bit 
    pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe' 
    #depends on your tesseract installation path
    custom_config = r'--oem 3 --psm 6'
    captcha = pytesseract.image_to_string(img2, config=custom_config)
    
    captcha.replace(" ", "").strip()

    print("Captcha printing " +captcha)
    print(len(captcha)-1)
    if(len(captcha)-1 != 6 ):
        return -1

    #finally input the result pages with required info.
    time.sleep(1)
    try:
        testbox.send_keys(usn)
        captchabox.send_keys(captcha) 
    except:
        return -1
    try:
        print(browser.current_url)
    except:
        return -1
    
    
    time.sleep(2)
    sub_codes = ["18CS51", "18CS52", "18CS53","18CS54","18CS55","18CS56","18CSL57","18CSL58","18CIV59"]
    rows = []

    for sub_code in sub_codes:
        subject = browser.find_element_by_xpath("//*[@id='dataPrint']//*[contains(text(),'"+sub_code+"')]//following::div[1]").text
        internal_marks = browser.find_element_by_xpath("//*[@id='dataPrint']//*[contains(text(),'"+sub_code+"')]//following::div[2]").text
        external_marks = browser.find_element_by_xpath("//*[@id='dataPrint']//*[contains(text(),'"+sub_code+"')]//following::div[3]").text
        total_marks = browser.find_element_by_xpath("//*[@id='dataPrint']//*[contains(text(),'"+sub_code+"')]//following::div[4]").text
        remarks = browser.find_element_by_xpath("//*[@id='dataPrint']//*[contains(text(),'"+sub_code+"')]//following::div[5]").text

        present_row_data={'Subject Code': sub_code,
                   'Subject Name': subject,
                   'Internal Marks': internal_marks,
                   'External Marks': external_marks,
                   'Total': total_marks,
                   'Remarks': remarks }
        rows.append(present_row_data)
    
    final_result_data = pd.DataFrame(rows)                              #import pandas as pd
    final_result_data.to_excel(r'vtu_result.xlsx',index=False)


    filepath=r"C:\Users\harsh\Desktop\result_analysis\pics\vtu_result.xlsx"
    wb=load_workbook(filepath)
    sheet=wb.active
    sub1_i = sheet["C2"]
    sub1_e = sheet["D2"]
    sub1_t = sheet["E2"]
    sub1_r = sheet["F2"]

    sub2_i = sheet["C3"]
    sub2_e = sheet["D3"]
    sub2_t = sheet["E3"]
    sub2_r = sheet["F3"]

    sub3_i = sheet["C4"]
    sub3_e = sheet["D4"]
    sub3_t = sheet["E4"]
    sub3_r = sheet["F4"]

    sub4_i = sheet["C5"]
    sub4_e = sheet["D5"]
    sub4_t = sheet["E5"]
    sub4_r = sheet["F5"]

    sub5_i = sheet["C6"]
    sub5_e = sheet["D6"]
    sub5_t = sheet["E6"]
    sub5_r = sheet["F6"]
    
    sub6_i = sheet["C7"]
    sub6_e = sheet["D7"]
    sub6_t = sheet["E7"]
    sub6_r = sheet["F7"]

    sub7_i = sheet["C8"]
    sub7_e = sheet["D8"]
    sub7_t = sheet["E8"]
    sub7_r = sheet["F8"]

    sub8_i = sheet["C9"]
    sub8_e = sheet["D9"]
    sub8_t = sheet["E9"]
    sub8_r = sheet["F9"]

    sub9_i = sheet["C10"]
    sub9_e = sheet["D10"]
    sub9_t = sheet["E10"]
    sub9_r = sheet["F10"]

    workbook = load_workbook(r"C:\Users\harsh\Desktop\result_analysis\pics\student_marks_list.xlsx")
    sheet1 = workbook.active
    sheet1["B" + str(ite)] = sub1_i.value
    sheet1["C" + str(ite)] = sub1_e.value
    sheet1["D" + str(ite)] = sub1_t.value
    sheet1["E" + str(ite)] = sub1_r.value

    sheet1["F" + str(ite)] = sub2_i.value
    sheet1["G" + str(ite)] = sub2_e.value
    sheet1["H" + str(ite)] = sub2_t.value
    sheet1["I" + str(ite)] = sub2_r.value

    sheet1["J" + str(ite)] = sub3_i.value
    sheet1["K" + str(ite)] = sub3_e.value
    sheet1["L" + str(ite)] = sub3_t.value
    sheet1["M" + str(ite)] = sub3_r.value

    sheet1["N" + str(ite)] = sub4_i.value
    sheet1["O" + str(ite)] = sub4_e.value
    sheet1["P" + str(ite)] = sub4_t.value
    sheet1["Q" + str(ite)] = sub4_r.value

    sheet1["R" + str(ite)] = sub5_i.value
    sheet1["S" + str(ite)] = sub5_e.value
    sheet1["T" + str(ite)] = sub5_t.value
    sheet1["U" + str(ite)] = sub5_r.value

    sheet1["V" + str(ite)] = sub6_i.value
    sheet1["W" + str(ite)] = sub6_e.value
    sheet1["X" + str(ite)] = sub6_t.value
    sheet1["Y" + str(ite)] = sub6_r.value

    sheet1["Z" + str(ite)] = sub7_i.value
    sheet1["AA" + str(ite)] = sub7_e.value
    sheet1["AB" + str(ite)] = sub7_t.value
    sheet1["AC" + str(ite)] = sub7_r.value

    sheet1["AD" + str(ite)] = sub8_i.value
    sheet1["AE" + str(ite)] = sub8_e.value
    sheet1["AF" + str(ite)] = sub8_t.value
    sheet1["AG" + str(ite)] = sub8_r.value

    sheet1["AH" + str(ite)] = sub9_i.value
    sheet1["AI" + str(ite)] = sub9_e.value
    sheet1["AJ" + str(ite)] = sub9_t.value
    sheet1["AK" + str(ite)] = sub9_r.value
    

    workbook.save(r"C:\Users\harsh\Desktop\result_analysis\pics\student_marks_list.xlsx")

    time.sleep(2)
    return ite
    
    

filepath=r"C:\Users\harsh\Desktop\result_analysis\pics\student_marks_list.xlsx"
wb=load_workbook(filepath)
sheet=wb.active

def main():
    ite=3
    print("START")
    while ite <= sheet.max_row:
        cell_obj = sheet.cell(row=ite, column=1)
        usn = cell_obj.value
        x = fillLoginpage(usn, ite)
        print("IN MAIN FUNC") #for testing
        print(ite) #for testing
        if(x == 1):
            print(x)
            ite = ite+1
            continue
        elif(x == -1):
            print("IN ekif block")
            print(ite)
            continue
        ite = ite+1

    
if __name__ == "__main__":
    main()



    
    





