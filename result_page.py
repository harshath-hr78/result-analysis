from distutils.log import error
from selenium import webdriver
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
from openpyxl import load_workbook,cell
import time
import os
import pyautogui
import cv2 as cv
import pytesseract
from bs4 import BeautifulSoup
import pandas as pd
import requests

# install all these as pip install filename, and pip install opencv-python.

option = webdriver.ChromeOptions()
option.add_argument("-incognito")
option.add_experimental_option("excludeSwitches", ['enable-automation'])
option.add_experimental_option("detach",True)

#add your chrome driver installation path
browser = webdriver.Chrome(executable_path=r'C:\Program Files (x86)\chromedriver.exe', options=option)

def fillLoginpage(usn,ite):

    browser.get(r"https://results.vtu.ac.in/FMEcbcs22/resultpage.php")

    #getting hold of usn and captcha input fields.

    testbox = browser.find_element_by_xpath("/html/body/div[2]/div[1]/div[2]/div/div[2]/form/div/div[2]/div[1]/div/input")
    captchabox = browser.find_element_by_xpath("/html/body/div[2]/div[1]/div[2]/div/div[2]/form/div/div[2]/div[2]/div[1]/input")

    #start with the image capta recognition procedure
    time.sleep(2)
    myScreenshot = pyautogui.screenshot(region=(670,510,230,110)) #region=(horizontal pos, vertical pos, vertical ratio, horizontal ratio)
    myScreenshot.save(r'C:\Users\harsh\Desktop\result_analysis\pics\screenshot.png') #change according to your dir.

    os.chdir(r"C:\Users\harsh\Desktop\result_analysis\pics")
    img = cv.imread('screenshot.png',0)
    ret,thresh = cv.threshold(img,103,150,cv.THRESH_TOZERO_INV)
    cv.imshow('Binary Threshold', thresh)
    # Using cv2.imwrite() method
    # Saving the image
    os.chdir(r'C:\Users\harsh\Desktop\result_analysis\pics')
    cv.imwrite("thresh_img.png", thresh)

    time.sleep(1)
    #os.system('"wsl tesseract thresh_img.jpg result"') #tesseract is ocr function for image to text
    img2 = cv.imread('thresh_img.png',0)
    #install tesseract from https://github.com/UB-Mannheim/tesseract/wiki choose 64-bit 
    pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe' 
    #depends on your tesseract installation path
    custom_config = r'--oem 3 --psm 6'
    captcha = pytesseract.image_to_string(img2, config=custom_config)
    
    captcha.replace(" ", "").strip()

    print("Captcha printing " +captcha)
    print(len(captcha)-1)
    if(len(captcha)-1 != 6 ):
        return -1

    #finally input the result pages with required info.
    time.sleep(1)
    try:
        testbox.send_keys(usn)
        captchabox.send_keys(captcha) 
    except:
        return -1
    try:
        print(browser.current_url)
    except:
        return -1
    
    
    time.sleep(2)
    sub_codes = ["18ME751", "18CS71", "18CS72","18CS744","18CS734","18CSL76","18CSP77"]
    rows = []

    for sub_code in sub_codes:
        subject = browser.find_element_by_xpath("//*[@id='dataPrint']//*[contains(text(),'"+sub_code+"')]//following::div[1]").text
        internal_marks = browser.find_element_by_xpath("//*[@id='dataPrint']//*[contains(text(),'"+sub_code+"')]//following::div[2]").text
        external_marks = browser.find_element_by_xpath("//*[@id='dataPrint']//*[contains(text(),'"+sub_code+"')]//following::div[3]").text
        total_marks = browser.find_element_by_xpath("//*[@id='dataPrint']//*[contains(text(),'"+sub_code+"')]//following::div[4]").text
        remarks = browser.find_element_by_xpath("//*[@id='dataPrint']//*[contains(text(),'"+sub_code+"')]//following::div[5]").text

        present_row_data={'Subject Code': sub_code,
                   'Subject Name': subject,
                   'Internal Marks': internal_marks,
                   'External Marks': external_marks,
                   'Total': total_marks,
                   'Remarks': remarks }
        rows.append(present_row_data)
    
    final_result_data = pd.DataFrame(rows)                              #import pandas as pd
    final_result_data.to_excel(r'vtu_result.xlsx',index=False)


    filepath=r"C:\Users\harsh\Desktop\result_analysis\pics\vtu_result.xlsx"
    wb=load_workbook(filepath)
    sheet=wb.active
    energy_i = sheet["C2"]
    energy_e = sheet["D2"]
    energy_t = sheet["E2"]

    ai_i = sheet["C3"]
    ai_e = sheet["D3"]
    ai_t = sheet["E3"]

    big_i = sheet["C4"]
    big_e = sheet["D4"]
    big_t = sheet["E4"]

    crypto_i = sheet["C5"]
    crypto_e = sheet["D5"]
    crypto_t = sheet["E5"]

    uid_i = sheet["C6"]
    uid_e = sheet["D6"]
    uid_t = sheet["E6"]
    
    ailab_i = sheet["C7"]
    ailab_e = sheet["D7"]
    ailab_t = sheet["E7"]

    project_i = sheet["C8"]
    project_e = sheet["D8"]
    project_t = sheet["E8"]

    workbook = load_workbook(r"C:\Users\harsh\Desktop\result_analysis\pics\student_marks_list.xlsx")
    sheet1 = workbook.active
    sheet1["B" + str(ite)] = energy_i.value
    sheet1["C" + str(ite)] = energy_e.value
    sheet1["D" + str(ite)] = energy_t.value

    sheet1["E" + str(ite)] = ai_i.value
    sheet1["F" + str(ite)] = ai_e.value
    sheet1["G" + str(ite)] = ai_t.value

    sheet1["H" + str(ite)] = big_i.value
    sheet1["I" + str(ite)] = big_e.value
    sheet1["J" + str(ite)] = big_t.value

    sheet1["K" + str(ite)] = crypto_i.value
    sheet1["L" + str(ite)] = crypto_e.value
    sheet1["M" + str(ite)] = crypto_t.value

    sheet1["N" + str(ite)] = uid_i.value
    sheet1["O" + str(ite)] = uid_e.value
    sheet1["P" + str(ite)] = uid_t.value

    sheet1["Q" + str(ite)] = ailab_i.value
    sheet1["R" + str(ite)] = ailab_e.value
    sheet1["S" + str(ite)] = ailab_t.value

    sheet1["T" + str(ite)] = project_i.value
    sheet1["U" + str(ite)] = project_e.value
    sheet1["V" + str(ite)] = project_t.value
    

    workbook.save(r"C:\Users\harsh\Desktop\result_analysis\pics\student_marks_list.xlsx")

    time.sleep(5)
    return ite
    
    

filepath=r"C:\Users\harsh\Desktop\result_analysis\pics\student_marks_list.xlsx"
wb=load_workbook(filepath)
sheet=wb.active

def main():
    ite=3
    print("START")
    while ite <= sheet.max_column:
        cell_obj = sheet.cell(row=ite, column=1)
        usn = cell_obj.value
        x = fillLoginpage(usn, ite)
        print("IN MAIN FUNC") #for testing
        print(ite) #for testing
        if(x == ite):
            print(x)
        elif(x == -1):
            print(x)
            continue
        ite = ite+1

    
if __name__ == "__main__":
    main()



    
    






    





    
    









    
    






